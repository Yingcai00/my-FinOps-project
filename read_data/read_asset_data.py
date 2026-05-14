import json
from openpyxl import load_workbook

# 读取Excel文件
workbook = load_workbook('资产数据_2026-04-13 0917.xlsx')
sheet = workbook.active

# 获取列名
columns = []
for cell in sheet[1]:
    columns.append(cell.value)

print("Excel列名:")
print(columns)

# 数据清洗和转换
asset_data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    row_dict = {}
    for i, value in enumerate(row):
        if i < len(columns):
            row_dict[columns[i]] = value
    
    asset = {
        'id': len(asset_data) + 1,
        'code': str(row_dict.get('资产编码', '')),
        'name': str(row_dict.get('资产名称', '')),
        'status': str(row_dict.get('设备状态', '')),
        'model': str(row_dict.get('资产型号', '')),
        'config': str(row_dict.get('资产配置', '')),
        'serialNumber': str(row_dict.get('设备序列号', '')),
        'categoryName': str(row_dict.get('设备分类', '')),
        'finopsCategory': str(row_dict.get('资产细分', '')),
        'finopsPurpose': str(row_dict.get('部署模式', '')),
        'unit': str(row_dict.get('单位', '')),
        'owner': str(row_dict.get('资产归属公司', ''))
    }
    asset_data.append(asset)

# 生成JavaScript数据
js_data = 'const assetData = ' + json.dumps(asset_data, ensure_ascii=False, indent=2) + ';'

# 保存到文件
with open('asset_data.js', 'w', encoding='utf-8') as f:
    f.write(js_data)

print(f"成功读取 {len(asset_data)} 条资产数据")
print("数据已保存到 asset_data.js 文件")
